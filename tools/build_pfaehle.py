# -*- coding: utf-8 -*-
"""Erzeugt uploads/pfaehle.json fuer die Gewinde-Inventur.

Quelle (Dropbox, READ ONLY): die vier ILF-Daten-Listen mit den Absteckpunkten
je Stuetze (s1..s4_rechts/_hoch) + die Belegung-Listen fuer den Tischtyp.
Ausgabe je Pfahl: [tischId, stuetze, lat, lon, typ]  (WGS84, 6 Nachkommastellen)

Pfahllaenge nach Tischtyp (Angabe V. Malt 31.07.2026):
  Typ E        -> 3.0 m
  Typ C und D  -> 2.5 m
  Typ A/B      -> in der Anzeige nur informativ (nicht Teil der Inventur-Frage)
"""
import io, json, os

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DROPBOX = (r"C:\Users\MaltVic\STRABAG SE Dropbox\CH-MX-EE-Projekte\EE-RK-240 AUSF"
           r"\EE-RKDO_2400024369_NalpSolar\07 Plaene-Gutachten-Statik"
           r"\01 Plane Ablage\Ausführungspläne\PV-Perimeter")

DATEN = {
    'PZ2025': DROPBOX + r"\2026\PZ2025\NalpSolar_51_LI_0301_A9_Daten-2025_260710.xlsx",
    'PZ1':    DROPBOX + r"\2026\PZ1\NalpSolar_51_LI_0305_A3_Daten-2026-PZ1_260710.xlsx",
    'PZ2':    DROPBOX + r"\2026\PZ2\NalpSolar_51_LI_0307_A2_Daten-26plus-PZ2_260520.xlsx",
    'PZ3':    DROPBOX + r"\2026\PZ3\NalpSolar_51_LI_0309_A1_Daten-2027-PZ3_260612.xlsx",
    'ALT2025': DROPBOX + r"\alt\NalpSolar_51_LI_0301_A5_Daten-2025_250603.xlsx",
}

def lv95_wgs84(E, N):
    y = (E - 2600000.0) / 1e6
    x = (N - 1200000.0) / 1e6
    lam = 2.6779094 + 4.728982 * y + 0.791484 * y * x + 0.1306 * y * x * x - 0.0436 * y ** 3
    phi = 16.9023892 + 3.238272 * x - 0.270978 * y * y - 0.002528 * x * x - 0.0447 * y * y * x - 0.0140 * x ** 3
    return round(phi * 100 / 36, 6), round(lam * 100 / 36, 6)

def typen_aus_website():
    """Tischtyp je ID aus der bereits erzeugten modultische_all.json (Feld 7)."""
    p = os.path.join(BASE, 'uploads', 'tables', 'modultische_all.json')
    typ = {}
    for t in json.load(io.open(p, encoding='utf-8')):
        if len(t) > 6 and t[6]:
            typ[str(t[0])] = str(t[6])
    return typ

def main():
    import openpyxl
    typen = typen_aus_website()
    pfaehle = {}
    for pz, pfad in DATEN.items():
        if not os.path.exists(pfad):
            print('  ! fehlt:', pz); continue
        wb = openpyxl.load_workbook(pfad, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        hdr = [str(c).strip().lower() if c is not None else '' for c in next(rows)]
        if 'id' not in hdr:
            print('  ! ohne id-Spalte:', pz); wb.close(); continue
        ii = hdr.index('id')
        n = 0
        for r in rows:
            if r[ii] is None:
                continue
            tid = str(r[ii]).strip()
            if tid.endswith('.0'):
                tid = tid[:-2]
            for s in ('s1', 's2', 's3', 's4'):
                ke, kn = s + '_rechts', s + '_hoch'
                if ke not in hdr or kn not in hdr:
                    continue
                E, N = r[hdr.index(ke)], r[hdr.index(kn)]
                if E is None or N is None:
                    continue
                try:
                    lat, lon = lv95_wgs84(float(E), float(N))
                except (TypeError, ValueError):
                    continue
                key = tid + '-' + s.upper()
                if key in pfaehle:          # aktuelle Listen zuerst -> nicht ueberschreiben
                    continue
                pfaehle[key] = [tid, s.upper(), lat, lon, typen.get(tid, '')]
                n += 1
        wb.close()
        print('  %-8s %5d Pfahlpunkte' % (pz, n))

    aus = sorted(pfaehle.values(), key=lambda p: (int(p[0]) if p[0].isdigit() else 0, p[1]))
    ziel = os.path.join(BASE, 'uploads', 'pfaehle.json')
    io.open(ziel, 'w', encoding='utf-8').write(
        '[\n' + ',\n'.join(json.dumps(p, ensure_ascii=False) for p in aus) + '\n]\n')
    print('-> %s  (%d Pfaehle, %d Tische)' % (ziel, len(aus), len(set(p[0] for p in aus))))
    ohne = sorted(set(p[0] for p in aus if not p[4]))
    if ohne:
        print('   ohne Typangabe: %d Tische (z.B. %s)' % (len(ohne), ', '.join(ohne[:8])))

if __name__ == '__main__':
    main()

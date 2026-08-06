# -*- coding: utf-8 -*-
r"""
build_pz3_karte.py – PZ3-Modultische in die 2D-Karte (karte.html) einbauen.

Quelle: PZ3-Belegungsliste (Dropbox, READ ONLY):
  NalpSolar_51_LI_0309_A1_Belegung-2027-PZ3_260612.xlsx
  (gleiche Quelle wie tools/build_baustand.py; bei neuer Revision Pfad anpassen
  und dieses Skript erneut laufen lassen – der pz3Data-Block wird ersetzt.)

Geometrie: LineString = Tisch-Laengsachse, Zentrum +/- tisch_laenge/2 in
Richtung tisch_ausrichtung+90 (az = Modul-Blickrichtung, Achse senkrecht dazu).
LV95 -> WGS84 via pyproj + konstanter Offset (dLon +0.000000063, dLat
-0.000005554), empirisch aus allen 1176 Endpunkten der bestehenden
pz1/pz2/pz2025-Layer bestimmt (Restfehler < 1 cm, 06.08.2026) – die
Original-Layer wurden mit einer anderen Transformation exportiert; ohne den
Offset saessen die PZ3-Tische 0.62 m verschoben neben den bestehenden.

Patcht in karte.html (nur beim ersten Lauf; danach wird pz3Data ersetzt):
  1. Datenblock  const pz3Data = {...};   vor  const pz1Layer = ...
  2. Layerzeile  const pz3Layer = addTableLayer(pz3Data, 'PZ3');
  3. Layer-Control-Eintrag 'PZ3 Modultische (2027)'
  4. Statuszeile (Modultisch-Zaehler)

Aufruf (aus Website_NalpSolar):  python tools/build_pz3_karte.py
Danach: sw.js-Cache-Version hochzaehlen + commit/push nicht vergessen.
"""
import json, math, os, re, sys

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PZ3_XLSX = (r"C:\Users\MaltVic\STRABAG SE Dropbox\CH-MX-EE-Projekte\EE-RK-240 AUSF"
            r"\EE-RKDO_2400024369_NalpSolar\07 Plaene-Gutachten-Statik"
            r"\01 Plane Ablage\Ausführungspläne\PV-Perimeter\2026"
            r"\PZ3\NalpSolar_51_LI_0309_A1_Belegung-2027-PZ3_260612.xlsx")
PZ3_QUELLE = os.path.basename(PZ3_XLSX)

# Empirischer Welt-Offset (siehe Docstring)
D_LON = +0.000000063
D_LAT = -0.000005554

SPALTEN = ['betriebs_id', 'id', 'zentrum_rechts', 'zentrum_hoch', 'zentrum_hoehe',
           'modul_hoehe', 'tisch_art', 'tisch_typ', 'tisch_ausrichtung',
           'tisch_laenge', 'baujahr', 'bauphase', 'sperrzone', 'gak', 'bemerkung']


def lese_pz3():
    import openpyxl
    wb = openpyxl.load_workbook(PZ3_XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    hdr = [str(c).strip().lower() if c is not None else '' for c in next(rows)]
    ix = {k: hdr.index(k) for k in SPALTEN}
    tische = []
    for r in rows:
        if r[ix['id']] is None:
            continue
        t = {k: r[ix[k]] for k in SPALTEN}
        try:
            for k in ('zentrum_rechts', 'zentrum_hoch', 'zentrum_hoehe',
                      'modul_hoehe', 'tisch_ausrichtung', 'tisch_laenge'):
                t[k] = float(t[k])
            t['id'] = int(t['id'])
        except (TypeError, ValueError):
            print('  ! Zeile mit ID %r unvollstaendig - uebersprungen' % (t['id'],))
            continue
        tische.append(t)
    wb.close()
    return tische


def baue_geojson(tische):
    from pyproj import Transformer
    tr = Transformer.from_crs(2056, 4326, always_xy=True)

    def w(E, N):
        lon, lat = tr.transform(E, N)
        return [round(lon + D_LON, 9), round(lat + D_LAT, 9)]

    feats = []
    for t in tische:
        az = math.radians(t['tisch_ausrichtung'] + 90.0)
        dE = math.sin(az) * t['tisch_laenge'] / 2.0
        dN = math.cos(az) * t['tisch_laenge'] / 2.0
        props = {k: t[k] for k in SPALTEN}
        props['betriebs_id'] = str(props['betriebs_id'] or '')
        props['bemerkung'] = str(props['bemerkung'] or '')
        props['sperrzone'] = bool(props['sperrzone'])
        props['baujahr'] = int(props['baujahr'])
        props['quelle'] = PZ3_QUELLE
        feats.append({'type': 'Feature', 'properties': props,
                      'geometry': {'type': 'LineString', 'coordinates': [
                          w(t['zentrum_rechts'] - dE, t['zentrum_hoch'] - dN),
                          w(t['zentrum_rechts'] + dE, t['zentrum_hoch'] + dN)]}})
    return {'type': 'FeatureCollection', 'name': 'PZ3_Belegung_2027', 'features': feats}


def patch_karte(fc):
    pfad = os.path.join(BASE, 'karte.html')
    html = open(pfad, encoding='utf-8').read()
    daten_js = 'const pz3Data = ' + json.dumps(fc, ensure_ascii=False,
                                              separators=(',', ':')) + ';\n\n'

    if 'const pz3Data = ' in html:
        # Bestehenden Datenblock ersetzen (Listen-Revision)
        html = re.sub(r'const pz3Data = \{.*?\};\n\n', daten_js, html, count=1, flags=re.DOTALL)
        print('pz3Data-Block ERSETZT (%d Tische).' % len(fc['features']))
    else:
        anker = [
            ("const pz1Layer = addTableLayer(pz1Data, 'PZ1');",
             daten_js + "const pz1Layer = addTableLayer(pz1Data, 'PZ1');"),
            ("const pz2025Layer = addTableLayer(pz2025Data, 'PZ2025');",
             "const pz2025Layer = addTableLayer(pz2025Data, 'PZ2025');\n"
             "const pz3Layer = addTableLayer(pz3Data, 'PZ3');"),
            ("      { name: 'PZ2025 Modultische', layer: pz2025Layer },",
             "      { name: 'PZ2025 Modultische', layer: pz2025Layer },\n"
             "      { name: 'PZ3 Modultische (2027)', layer: pz3Layer },"),
            ("(pz1Data.features.length + pz2Data.features.length + pz2025Data.features.length)",
             "(pz1Data.features.length + pz2Data.features.length + pz2025Data.features.length + pz3Data.features.length)"),
        ]
        for alt, neu in anker:
            n = html.count(alt)
            if n != 1:
                print('ABBRUCH: Anker %d-mal gefunden statt 1-mal:\n%s' % (n, alt))
                sys.exit(1)
        for alt, neu in anker:
            html = html.replace(alt, neu, 1)
        print('pz3Data NEU eingebaut (%d Tische) + Layer/Control/Statuszeile.' % len(fc['features']))

    open(pfad, 'w', encoding='utf-8', newline='').write(html)

    # Verifikation: zurueckparsen
    html2 = open(pfad, encoding='utf-8').read()
    m = re.search(r'const pz3Data = (\{.*?\});\n', html2, re.DOTALL)
    fc2 = json.loads(m.group(1))
    assert len(fc2['features']) == len(fc['features'])
    typen = {}
    for f in fc2['features']:
        typen[f['properties']['tisch_typ']] = typen.get(f['properties']['tisch_typ'], 0) + 1
    print('Verifiziert: %d PZ3-Features, Typen: %s' % (
        len(fc2['features']), ', '.join('%s=%d' % kv for kv in sorted(typen.items()))))


if __name__ == '__main__':
    tische = lese_pz3()
    print('PZ3-Belegungsliste: %d Tische (%s)' % (len(tische), PZ3_QUELLE))
    patch_karte(baue_geojson(tische))

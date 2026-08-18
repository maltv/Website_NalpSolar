# -*- coding: utf-8 -*-
"""Bestellbatch(es) in assets/js/nalp-passstueck-daten.js nachtragen.

Aufruf (aus Website_NalpSolar):
    python tools/bestellbatch_nachtragen.py <Bestellbatch.xlsx> [weitere.xlsx ...]
    python tools/bestellbatch_nachtragen.py --pruefen <...>     # nur anzeigen, nichts schreiben

Quelle ist immer die an Mauchle gesendete Bestell-xlsx (Blatt 1, ab Zeile 5):
    A Bestellbatch | B ID Tisch | C Tisch Typ | D ID Fuss | E ABS_PS_T/B [mm] | F Bauteilart
Bestelldatum steht in Zeile 1, Spalte F.

Datensatz je Position (Reihenfolge wie bisher im JS):
    [tisch, stuetze, laenge_mm, batch, datum, bauteilart, gueltig]

Regel "gueltig": kommt dieselbe Position (Tisch+Stuetze) in einem neu eingelesenen
Batch nochmals vor, ist nur die NEUESTE Bestellung gueltig=1; alle aelteren Zeilen
dazu werden auf 0 (=ersetzt) gesetzt. Flags von Positionen, die der neue Batch NICHT
enthaelt, bleiben unangetastet (die Batch-5/6-Korrektur von 06/2026 ist kuratiert).
Der Lauf ist idempotent: ein bereits eingetragener Batch wird ersetzt, nicht doppelt.

Sonderbestellungen mit nicht-numerischem Batch (z. B. "RES") kommen aus
tools/reserve_nachtragen.py; dieses Skript laesst sie unangetastet stehen.
"""
import json
import re
import shutil
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl

JS = Path(__file__).resolve().parent.parent / "assets" / "js" / "nalp-passstueck-daten.js"


def bkey(b):
    """Sortier-/Vergleichsschluessel fuer Batchnummern. Seit der Sonderbestellung
    RESERVE (tools/reserve_nachtragen.py) stehen neben Zahlen auch Kuerzel wie
    "RES" in der Liste - ohne diesen Helfer vergleicht Python int mit str."""
    return (0, b, "") if isinstance(b, int) else (1, 0, str(b))


def batch_lesen(pfad):
    """Bestell-xlsx -> (batchnr, bestelldatum ISO, [positionen])."""
    wb = openpyxl.load_workbook(pfad, data_only=True)
    ws = wb.worksheets[0]
    kopf = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    bestellt = kopf[5]
    if isinstance(bestellt, (datetime, date)):
        datum = bestellt.strftime("%Y-%m-%d")
    else:
        raise SystemExit("Bestelldatum (Zeile 1, Spalte F) fehlt in %s" % pfad)

    positionen, batches = [], set()
    for r in ws.iter_rows(min_row=5, values_only=True):
        if not r[1]:
            continue
        batches.add(int(str(r[0]).strip()))
        tisch = int(str(r[1]).strip())
        m = re.search(r"S\s*([1-4])\s*$", str(r[3]).strip(), re.I)
        if not m:
            raise SystemExit("ID Fuss unlesbar: %r (Tisch %s)" % (r[3], tisch))
        stuetze = "S" + m.group(1)
        laenge = float(r[4])
        art = str(r[5]).strip() if r[5] else "Normal"
        positionen.append([tisch, stuetze, laenge, None, datum, art, 1])

    if len(batches) != 1:
        raise SystemExit("Datei enthaelt mehrere Batchnummern: %s" % sorted(batches))
    nr = batches.pop()
    for p in positionen:
        p[3] = nr
    return nr, datum, positionen


def gueltig_nachziehen(bestellungen, betroffen):
    """Nur fuer die Positionen der neu eingelesenen Batches aufraeumen:
    neueste Bestellung gueltig, alle aelteren Zeilen dazu ersetzt (0).
    Alles andere bleibt so, wie es der urspruengliche Generator gesetzt hat."""
    neueste = {}
    for p in bestellungen:
        k = (int(p[0]), p[1])
        if k not in betroffen:
            continue
        rang = (p[4], bkey(p[3]))                  # Datum, dann Batchnummer
        if k not in neueste or rang > neueste[k]:
            neueste[k] = rang
    ersetzt = []
    for p in bestellungen:
        k = (int(p[0]), p[1])
        if k not in betroffen:
            continue
        soll = 1 if (p[4], bkey(p[3])) == neueste[k] else 0
        if soll != p[6]:
            ersetzt.append((p[3], k[0], k[1]))
        p[6] = soll
    return ersetzt


def main(argv):
    nur_pruefen = "--pruefen" in argv
    dateien = [a for a in argv if not a.startswith("--")]
    if not dateien:
        raise SystemExit(__doc__)

    text = JS.read_text(encoding="utf-8")
    kopf, _, rest = text.partition("window.NALP_PS_DATEN =")
    daten = json.loads(rest[: rest.rindex("}") + 1])
    best = daten["bestellungen"]
    vorher = len(best)

    betroffen = set()
    for f in dateien:
        nr, datum, pos = batch_lesen(f)
        alt = [p for p in best if p[3] == nr]
        best[:] = [p for p in best if p[3] != nr]
        best.extend(pos)
        betroffen |= {(int(p[0]), p[1]) for p in pos}
        print("Batch %03d  bestellt %s  %3d Positionen  %2d Tische%s"
              % (nr, datum, len(pos), len({p[0] for p in pos}),
                 "  (ersetzt %d vorhandene)" % len(alt) if alt else ""))

    best.sort(key=lambda p: (bkey(p[3]), int(p[0]), p[1]))
    geaendert = gueltig_nachziehen(best, betroffen)
    nr_alle = sorted({p[3] for p in best if isinstance(p[3], int)})
    sonder = sorted({str(p[3]) for p in best if not isinstance(p[3], int)})
    print("Positionen gesamt: %d -> %d | gueltig: %d | Batches: %d-%d%s"
          % (vorher, len(best), sum(1 for p in best if p[6] == 1), nr_alle[0], nr_alle[-1],
             " + " + ", ".join(sonder) if sonder else ""))
    if geaendert:
        je_batch = {}
        for b, t, s in geaendert:
            je_batch[b] = je_batch.get(b, 0) + 1
        print("Neu als 'ersetzt' markiert (spaeter nochmals bestellt): %d Positionen aus Batch %s"
              % (len(geaendert), ", ".join("%d (%dx)" % (b, n) for b, n in sorted(je_batch.items()))))

    if nur_pruefen:
        print("--pruefen: nichts geschrieben.")
        return

    daten["meta"]["bestellungen_quelle"] = (
        "Mauchle-Bestellbatches NR %d-%d (je Batch die an Mauchle gesendete xlsx; "
        "Batch 5+6 = Korrektur-Regenerierung 08.06.2026, Batch 11-15 nachgetragen "
        "%s aus den Bestellmails). Felder: tisch, stuetze, laenge_mm, batch, datum, "
        "bauteilart, gueltig(1)/ersetzt(0) - ersetzt = dieselbe Position wurde spaeter "
        "neu bestellt. ACHTUNG: belegt nur die BESTELLUNG (Basis = damaliger SOLL-Stand), "
        "nicht die Lieferung." % (nr_alle[0], nr_alle[-1], date.today().strftime("%d.%m.%Y"))
    )
    if sonder:
        daten["meta"]["bestellungen_quelle"] += (
            " Dazu Sonderbestellung(en) %s - siehe Block `reserve` "
            "(tools/reserve_nachtragen.py)." % ", ".join(sonder)
        )

    shutil.copy2(JS, JS.with_suffix(".js.bak"))
    neu = kopf + "window.NALP_PS_DATEN = " + json.dumps(daten, ensure_ascii=True,
                                                        separators=(",", ":")) + ";\n"
    JS.write_text(neu, encoding="utf-8")
    print("geschrieben: %s  (Sicherung: %s)" % (JS.name, JS.with_suffix(".js.bak").name))


if __name__ == "__main__":
    main(sys.argv[1:])

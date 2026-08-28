# -*- coding: utf-8 -*-
"""
build_baustand.py – erzeugt die beiden Datendateien fuer den 3D-Baustand:

  uploads/tables/modultische_all.json   [id,E,N,Z,az,len,typ] aller Tische
      Quelle: Dropbox (READ ONLY!) – Belegung-Excels der 4 Planungszonen:
      je Planungszone die HOECHSTE Revision im Ordner (Spalten zentrum_*, tisch_ausrichtung,
      tisch_laenge, tisch_typ). Fehlende IDs werden aus den bestehenden
      uploads/tables/modultische(.status).json ergaenzt (identisches Schema,
      Kalibrierung 09.07.2026: Abweichung 0.000).

  uploads/tables/baustand.json          { stand, quelle, tische:{id:{s,pk,mt}} }
      Quelle: Nalpi-Export "export-tableId-*.csv" (Victor legt ihn in
      _ZWISCHENABLAGE ab). Nur Tische MIT Datum "Primärkonstruktion montiert".
      s = 'mt' wenn auch "Modulträger montiert", sonst 'pk'.
      vs = Datum "Modultisch verschraubt", falls in Nalpi gesetzt.

Aufruf (aus Website_NalpSolar):
  python tools/build_baustand.py "<pfad zu export-tableId-JJJJ-MM-TT.csv>"

Bei neuem Nalpi-Export einfach erneut laufen lassen (nur baustand.json aendert
sich; die Koordinaten werden nur neu gebaut, wenn sich die Excels aendern).
"""
import csv, glob, io, json, os, re, sys

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DROPBOX = (r"C:\Users\MaltVic\STRABAG SE Dropbox\CH-MX-EE-Projekte\EE-RK-240 AUSF"
           r"\EE-RKDO_2400024369_NalpSolar\07 Plaene-Gutachten-Statik"
           r"\01 Plane Ablage\Ausführungspläne\PV-Perimeter\2026")

# Die Revision im Dateinamen wechselt (ILF schickt A3 -> A4 -> ...), und die
# alte Datei wandert dabei in den Unterordner "Alt". Ein fest verdrahteter
# Dateiname liess den Lauf am 27.08.2026 zwei Tage lang stillschweigend
# abstuerzen – der 3D-Baustand blieb auf dem 26.08. stehen. Darum wird jetzt je
# Ordner die HOECHSTE Revision gesucht (nur im Ordner selbst, nie in "Alt"),
# und ein Revisionswechsel wird gemeldet.
BELEGUNG_MUSTER = {
    'PZ2025': (r"\PZ2025", "NalpSolar_51_LI_0301_A*_Belegung-2025_*.xlsx"),
    'PZ1':    (r"\PZ1",    "NalpSolar_51_LI_0305_A*_Belegung-2026-PZ1_*.xlsx"),
    'PZ2':    (r"\PZ2",    "NalpSolar_51_LI_0307_A*_Belegung-26plus-PZ2_*.xlsx"),
    'PZ3':    (r"\PZ3",    "NalpSolar_51_LI_0309_A*_Belegung-2027-PZ3_*.xlsx"),
}
# Fuellquelle NUR fuer IDs, die oben fehlen (Serie 2025, gebaut Herbst 2025;
# steht in keiner der 4 aktuellen Listen mehr):
ALT2025 = os.path.dirname(DROPBOX) + r"\alt\NalpSolar_51_LI_0301_A5_Daten-2025_Belegung_250603.xlsx"
# Merkzettel ausserhalb des Web-Repos: welche Revision galt beim letzten Lauf
MERK = os.path.join(os.path.dirname(BASE), 'Tools', 'belegung_revisionen.json')


def revision(pfad):
    """'..._A4_...' -> 4 (fuer den Vergleich der Revisionen)."""
    m = re.search(r'_A(\d+)_', os.path.basename(pfad))
    return int(m.group(1)) if m else -1


def belegung_dateien():
    """{PZ: Pfad} der jeweils hoechsten Revision + Hinweis bei Wechsel."""
    try:
        frueher = json.load(io.open(MERK, encoding='utf-8'))
    except Exception:
        frueher = {}
    gewaehlt, jetzt = {}, {}
    for pz, (ordner, muster) in BELEGUNG_MUSTER.items():
        treffer = glob.glob(os.path.join(DROPBOX + ordner, muster))
        if not treffer:
            sys.exit('KEINE Belegungsliste gefunden fuer %s: %s\\%s'
                     % (pz, DROPBOX + ordner, muster))
        treffer.sort(key=lambda p: (revision(p), os.path.getmtime(p)))
        pfad = treffer[-1]
        gewaehlt[pz] = pfad
        jetzt[pz] = os.path.basename(pfad)
        alt = frueher.get(pz)
        if alt and alt != jetzt[pz]:
            print('  ! %s: NEUE Revision – vorher %s, jetzt %s' % (pz, alt, jetzt[pz]))
    gewaehlt['ALT2025'] = ALT2025
    try:
        os.makedirs(os.path.dirname(MERK), exist_ok=True)
        json.dump(jetzt, io.open(MERK, 'w', encoding='utf-8'), ensure_ascii=False, indent=1)
    except Exception as e:
        print('  ! Merkzettel nicht geschrieben: %s' % e)
    return gewaehlt


def lese_belegung():
    import openpyxl
    alle = {}
    for pz, pfad in belegung_dateien().items():
        wb = openpyxl.load_workbook(pfad, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        hdr = [str(c).strip().lower() if c is not None else '' for c in next(rows)]
        ix = {k: hdr.index(k) for k in ('id', 'zentrum_rechts', 'zentrum_hoch',
             'zentrum_hoehe', 'tisch_ausrichtung', 'tisch_laenge', 'tisch_typ', 'baujahr')}
        # Alte A5-Belegung (Serie 2025) hat kein modul_hoehe -> Standard
        # zentrum + 4.20 m (belegt: alle ILF-Listen 2026 exakt 4.2; Nalpi
        # TABLE_DATA H_SOLL-COORDZ 637x exakt 4.2, Auswertung 23.07.2026).
        hat_mh = 'modul_hoehe' in hdr
        if hat_mh:
            ix['modul_hoehe'] = hdr.index('modul_hoehe')
        n = 0
        for r in rows:
            if r[ix['id']] is None:
                continue
            tid = str(r[ix['id']]).strip()
            if pz == 'ALT2025' and tid in alle:
                continue          # Fuellquelle ueberschreibt nie aktuelle Listen
            try:
                # modul_hoehe = Unterkante Modulträger am Auflagerkopf (Bezugspunkt-
                # Skizze 2026\Bezugspunkt_Modul_Hoehe.png); in A9 tischweise korrigiert
                if hat_mh:
                    mh = round(float(r[ix['modul_hoehe']]), 3)
                else:
                    mh = round(float(r[ix['zentrum_hoehe']]) + 4.2, 3)
                rec = [tid,
                       round(float(r[ix['zentrum_rechts']]), 3),
                       round(float(r[ix['zentrum_hoch']]), 3),
                       round(float(r[ix['zentrum_hoehe']]), 3),
                       round(float(r[ix['tisch_ausrichtung']]), 2),
                       round(float(r[ix['tisch_laenge']]), 3),
                       str(r[ix['tisch_typ']] or '') + '_' + str(r[ix['baujahr']] or ''),
                       mh]
            except (TypeError, ValueError):
                print('  ! %s: Zeile mit ID %s unvollstaendig - uebersprungen' % (pz, tid))
                continue
            alle[tid] = rec
            n += 1
        wb.close()
        print('  %s: %d Tische  (%s)' % (pz, n, os.path.basename(pfad)))
    return alle

def ergaenze_bestehende(alle):
    """IDs, die nicht in den Belegung-Excels stehen, aus den vorhandenen
    3D-Dateien uebernehmen (Schema identisch, ohne Typ)."""
    for name in ('modultische.json', 'modultische_status.json'):
        p = os.path.join(BASE, 'uploads', 'tables', name)
        if not os.path.exists(p):
            continue
        extra = 0
        for t in json.load(io.open(p, encoding='utf-8')):
            tid = str(t[0]).strip()
            if tid not in alle:
                alle[tid] = [tid, t[1], t[2], t[3], t[4], t[5], '']
                extra += 1
        print('  +%d aus %s' % (extra, name))
    return alle

def lese_nalpi(csv_pfad):
    st = {}
    stati = {}
    with io.open(csv_pfad, encoding='utf-8-sig', newline='') as fh:
        for r in csv.DictReader(fh):
            tid = r['Nr.'].strip()
            pk = r.get('Primärkonstruktion montiert', '').strip()
            mt = r.get('Modulträger montiert', '').strip()
            # Spalte gibt es erst seit 18.08.2026 – aeltere CSV haben sie nicht
            vs = r.get('Modultisch verschraubt', '').strip()
            stati[r.get('Baustatus', '?')] = stati.get(r.get('Baustatus', '?'), 0) + 1
            if pk in ('', '—', '-'):
                continue
            e = {'s': 'mt' if mt not in ('', '—', '-') else 'pk', 'pk': pk}
            if e['s'] == 'mt':
                e['mt'] = mt
            if vs not in ('', '—', '-'):
                e['vs'] = vs          # Modultisch verschraubt (Nalpi)
            st[tid] = e
    return st, stati

def main():
    csv_pfad = sys.argv[1] if len(sys.argv) > 1 else None

    print('Belegung-Excels (Dropbox, nur lesen):')
    alle = ergaenze_bestehende(lese_belegung())
    print('Total Koordinaten: %d Tische' % len(alle))

    out1 = os.path.join(BASE, 'uploads', 'tables', 'modultische_all.json')
    vals = sorted(alle.values(), key=lambda t: int(t[0]) if t[0].isdigit() else 0)
    io.open(out1, 'w', encoding='utf-8').write(
        '[\n' + ',\n'.join(json.dumps(v, ensure_ascii=False) for v in vals) + '\n]\n')
    print('-> %s (%d)' % (out1, len(vals)))

    if not csv_pfad or not os.path.exists(csv_pfad):
        print('Kein Nalpi-Export uebergeben -> baustand.json bleibt unveraendert.')
        return
    stand = os.path.basename(csv_pfad).replace('export-tableId-', '').replace('.csv', '')

    st, stati = lese_nalpi(csv_pfad)
    n_pk = sum(1 for v in st.values() if v['s'] == 'pk')
    n_mt = sum(1 for v in st.values() if v['s'] == 'mt')
    print('Baustand %s: %d gestellt (%d nur PK, %d mit Modultraeger)' % (stand, len(st), n_pk, n_mt))
    fehlen = sorted([t for t in st if t not in alle], key=lambda x: int(x) if x.isdigit() else 0)
    if fehlen:
        print('OHNE Koordinaten (werden im 3D NICHT angezeigt): %d -> %s' % (len(fehlen), ','.join(fehlen)))

    out2 = os.path.join(BASE, 'uploads', 'tables', 'baustand.json')
    io.open(out2, 'w', encoding='utf-8').write(json.dumps(
        {'stand': stand, 'quelle': os.path.basename(csv_pfad), 'tische': st},
        ensure_ascii=False, indent=0))
    print('-> %s' % out2)

if __name__ == '__main__':
    main()

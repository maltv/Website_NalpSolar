# -*- coding: utf-8 -*-
"""Sonderbestellung RESERVE in assets/js/nalp-passstueck-daten.js nachtragen.

Aufruf (aus Website_NalpSolar):
    python tools/reserve_nachtragen.py <RESERVE.xlsx> [--pruefen]

Warum ein eigenes Skript: `bestellbatch_nachtragen.py` erwartet numerische
Batchnummern und Tisch-IDs. Die Sonderbestellung hat Batch "RES" und fuer die
Rohlinge Platzhalter-IDs ("RES-D2026"), die zu keinem Tisch gehoeren.

Die Datei enthaelt zweierlei und wird entsprechend getrennt abgelegt:

1. ROHLINGE (ID Tisch beginnt mit "RES-") -> neuer Block `reserve`.
   RRK-Hohlprofil MIT Lasche(n), OHNE Fussplatte, auf Maximallaenge des Typs.
   Zuschnitt fussseitig und Fussplatte anschweissen macht STRABAG selbst.
   Damit laesst sich ein zu kurz/falsch geliefertes Passstueck ersetzen, OHNE
   ein Stueck von einem anderen Tisch abzuziehen (Praezedenz: 1076_S4 wurde als
   Spender fuer Tisch 370 verbaut und fehlt seither an seiner Originalposition).

2. ECHTE NACHBESTELLUNGEN (numerische ID Tisch) -> in `bestellungen`, Batch "RES".
   Normalausfuehrung, komplettes Passstueck mit Fussplatte.
"""
import json
import re
import shutil
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl

JS = Path(__file__).resolve().parent.parent / "assets" / "js" / "nalp-passstueck-daten.js"
BATCH = "RES"
# Zusagen aus dem Mailthread "Passstueck Sonderbestellung RESERVE - NalpSolar"
VORGEZOGEN = {"1076_S4": "2026-08-20"}   # Maeder 17.08.2026 15:27 + 16:48


def lesen(pfad):
    wb = openpyxl.load_workbook(pfad, data_only=True)
    ws = wb.worksheets[0]
    kopf = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if not isinstance(kopf[5], (datetime, date)):
        raise SystemExit("Bestelldatum (Zeile 1, Spalte F) fehlt in %s" % pfad)
    bestellt = kopf[5].strftime("%Y-%m-%d")
    liefer = kopf[9].strftime("%Y-%m-%d") if isinstance(kopf[9], (datetime, date)) else None

    rohlinge, nachbestellt = {}, []
    for r in ws.iter_rows(min_row=5, values_only=True):
        if not r[1]:
            continue
        tisch, typ = str(r[1]).strip(), str(r[2]).strip()
        m = re.search(r"S\s*([1-4])\s*$", str(r[3]).strip(), re.I)
        if not m:
            raise SystemExit("ID Fuss unlesbar: %r" % (r[3],))
        st = "S" + m.group(1)
        laenge = round(float(r[4]))
        art = str(r[5]).strip() if r[5] else "Normal"
        if tisch.upper().startswith("RES-"):
            # Tal = S1/S4, Berg = S2/S3; je Lage sind es 2 Stueck gleicher Laenge
            lage = "T" if st in ("S1", "S4") else "B"
            k = (typ, lage, laenge)
            rohlinge[k] = rohlinge.get(k, 0) + 1
        else:
            nachbestellt.append([int(tisch), st, float(laenge), BATCH, bestellt, art, 1])
    stueck = [[t, l, ln, n] for (t, l, ln), n in
              sorted(rohlinge.items(), key=lambda kv: (kv[0][0], kv[0][1] != "T"))]
    return bestellt, liefer, stueck, nachbestellt


def main(argv):
    nur_pruefen = "--pruefen" in argv
    dateien = [a for a in argv if not a.startswith("--")]
    if not dateien:
        raise SystemExit(__doc__)
    bestellt, liefer, stueck, nachbestellt = lesen(dateien[0])

    text = JS.read_text(encoding="utf-8")
    kopf, _, rest = text.partition("window.NALP_PS_DATEN =")
    daten = json.loads(rest[: rest.rindex("}") + 1])
    best = daten["bestellungen"]

    print("Sonderbestellung RESERVE, bestellt %s, Liefertermin %s" % (bestellt, liefer))
    print("  Rohlinge: %d Stueck in %d Positionen" % (sum(s[3] for s in stueck), len(stueck)))
    for typ, lage, ln, n in stueck:
        print("    %-8s %s  %5d mm  x%d" % (typ, "Tal " if lage == "T" else "Berg", ln, n))
    print("  Nachbestellungen (echte Tische, in `bestellungen`): %d" % len(nachbestellt))
    for p in nachbestellt:
        vor = VORGEZOGEN.get("%d_%s" % (p[0], p[1]))
        frueher = [b for b in best if int(b[0]) == p[0] and b[1] == p[1] and b[3] != BATCH]
        print("    %d_%s  %d mm%s%s" % (
            p[0], p[1], round(p[2]),
            "  (vorgezogen auf %s)" % vor if vor else "",
            "  | vorher Batch %s" % ", ".join(str(b[3]) for b in frueher) if frueher else ""))

    # idempotent: vorhandenen RES-Stand ersetzen, nicht doppeln
    best[:] = [p for p in best if p[3] != BATCH]
    best.extend(nachbestellt)
    best.sort(key=lambda p: (str(p[3]).zfill(3), int(p[0]), p[1]))

    daten["reserve"] = {
        "batch": BATCH,
        "bestellt": bestellt,
        "liefertermin": liefer,
        "vorgezogen": VORGEZOGEN,
        "quelle": r"Bestellungen\20260806_Stahlbau_Bestellbatch_RESERVE_NalpSolar.xlsx",
        "spalten": ["typ", "lage", "laenge_mm", "anzahl"],
        "ausfuehrung": "RRK-Hohlprofil MIT angeschweisster Lasche(n), OHNE Fussplatte. "
                       "Zuschnitt fussseitig auf das benötigte Mass und Anschweissen der "
                       "Fussplatte erfolgt bauseits durch STRABAG.",
        "zweck": "Ersatz, wenn ein Passstück zu kurz/zu lang geliefert oder falsch berechnet "
                 "ist – ohne einem anderen Tisch sein Stück wegzunehmen.",
        "hinweis": "Länge je Position = längste bisher bestellte Passstücklänge des Typs "
                   "und der Stützenart (Batches 1–14 inkl. Batch-5-Korrektur + Lieferungen "
                   "Serie 2025), aufgerundet. Ein Rohling passt für jede Länge BIS zu diesem "
                   "Mass. Typ A_2026 nicht enthalten (nie bestellt, keine Referenzlänge); "
                   "Typ C existiert in Serie 2025 nicht. Lage T = Tal (S1/S4), B = Berg (S2/S3).",
        "stueck": stueck,
    }

    if nur_pruefen:
        print("--pruefen: nichts geschrieben.")
        return
    shutil.copy2(JS, JS.with_suffix(".js.bak"))
    neu = kopf + "window.NALP_PS_DATEN = " + json.dumps(daten, ensure_ascii=True,
                                                        separators=(",", ":")) + ";\n"
    JS.write_text(neu, encoding="utf-8")
    print("geschrieben: %s  (Sicherung: %s)" % (JS.name, JS.with_suffix(".js.bak").name))


if __name__ == "__main__":
    main(sys.argv[1:])
